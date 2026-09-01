import sys
import os
import re
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QLineEdit, QFileDialog, QMessageBox,
    QProgressBar, QFrame, QGraphicsDropShadowEffect, QScrollArea,
    QSizePolicy
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal
from PyQt6.QtGui import QFont, QColor, QPalette, QKeySequence, QShortcut, QPixmap
import tempfile
import shutil


# ─── Fix for Permission Issues ───────────────────────────────────────────────
def ensure_directory_permission(path):
    try:
        if not os.path.exists(path):
            os.makedirs(path, mode=0o777)
        test_file = os.path.join(path, "test_write.tmp")
        with open(test_file, 'w') as f:
            f.write("test")
        os.remove(test_file)
        return True
    except:
        return False


def get_safe_output_directory():
    safe_paths = [
        os.path.join(os.path.expanduser("~"), "Desktop"),
        os.path.join(os.path.expanduser("~"), "Documents"),
        tempfile.gettempdir(),
        os.getcwd(),
    ]
    for path in safe_paths:
        if ensure_directory_permission(path):
            return path
    return tempfile.mkdtemp()


# ─── Arabic Text Normalizer ──────────────────────────────────────────────────
def normalize_arabic(text: str) -> str:
    if not text:
        return ""
    t = str(text).strip()
    t = t.replace(" ", "").replace("-", "").replace("_", "")
    t = re.sub(r'[إأآا]', 'ا', t)
    t = re.sub(r'[يى]', 'ى', t)
    t = re.sub(r'[هة]', 'ه', t)
    t = t.replace('ـ', '')
    if t.startswith('ال'):
        t_no_al = t[2:]
    else:
        t_no_al = t
    return t_no_al or t


def find_col(df, candidates):
    norm_cols = {col: normalize_arabic(col) for col in df.columns}

    # 1️⃣ تطابق تام
    for cand in candidates:
        nc = normalize_arabic(cand)
        for orig, normed in norm_cols.items():
            if nc == normed:
                return orig

    # 2️⃣ تطابق يحتوي الكلمة (بس مش العكس)
    for cand in candidates:
        nc = normalize_arabic(cand)
        for orig, normed in norm_cols.items():
            if nc in normed:
                return orig

    return None


def normalize_id(val):
    if pd.isna(val):
        return ""
    s = str(val).strip().replace(".0", "").replace(" ", "")
    return ''.join(c for c in s if c.isdigit())


def format_date_only(val):
    if pd.isna(val):
        return ""
    s = str(val).strip()
    if ' ' in s:
        s = s.split(' ')[0]
    return s


def parse_date_flexible(val):
    """Parse a discharge / calendar date. Returns datetime or None."""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    raw = str(val).strip()
    if not raw or raw.lower() in ("nan", "none", "nat"):
        return None
    raw = raw.split()[0]
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d", "%d.%m.%Y"):
        try:
            return datetime.strptime(raw, fmt)
        except ValueError:
            continue
    d = pd.to_datetime(val, dayfirst=True, errors="coerce")
    if pd.notna(d):
        return d.to_pydatetime()
    return None


def add_months(dt: datetime, months: int) -> datetime:
    m0 = dt.month - 1 + months
    year = dt.year + m0 // 12
    month = m0 % 12 + 1
    if month == 12:
        nxt = datetime(year + 1, 1, 1)
    else:
        nxt = datetime(year, month + 1, 1)
    last_day = (nxt - pd.Timedelta(days=1)).day
    day = min(dt.day, int(last_day))
    return datetime(year, month, day)


def names_equivalent(a, b):
    """Compare two Arabic names after normalization. True / False / None."""
    na = normalize_arabic(a)
    nb = normalize_arabic(b)
    if not na or not nb:
        return None
    if na == nb:
        return True
    if na in nb or nb in na:
        return True
    return False


def classify_person_status(status_raw: str):
    """
    حالة الفرد:
      موجود     → allow
      مجلس طبي  → allow (exception)
      خارج      → deny
      else      → hard case
    """
    t = normalize_arabic(status_raw)
    raw = str(status_raw).strip() if status_raw is not None and not (isinstance(status_raw, float) and pd.isna(status_raw)) else ""
    if raw.lower() == "nan":
        raw = ""
    if not t:
        return "hard", raw
    if "مجلسطبي" in t or t == "مجلس":
        return "allow", raw or "مجلس طبي"
    if "موجود" in t:
        return "allow", raw
    if "خارج" in t:
        return "deny", raw
    return "hard", raw


def _discharge_ineligibility(discharge_str: str, payment_month_str: str):
    """
    Not eligible if discharged on day 1 of the payment month, or if the
    reference date (1st of payment month, else today) falls inside
    [discharge_date, discharge_date + 3 months].

    Returns (not_eligible: bool, reason: str, parsed datetime or None).
    """
    if not discharge_str or str(discharge_str).strip() in ("", "nan", "None"):
        return False, "", None

    d = parse_date_flexible(discharge_str)
    if d is None:
        return False, "HARD_DATE", None

    d = datetime(d.year, d.month, d.day)

    if payment_month_str and str(payment_month_str).strip():
        parts = str(payment_month_str).strip().split("-")
        if len(parts) != 2:
            return False, "HARD_DATE", d
        try:
            ref = datetime(int(parts[0]), int(parts[1]), 1)
        except ValueError:
            return False, "HARD_DATE", d
    else:
        today = datetime.now()
        ref = datetime(today.year, today.month, 1)

    window_end = add_months(d, 3)
    first_of_pay = (d.year == ref.year and d.month == ref.month and d.day == 1)
    in_3_month_window = d <= ref <= window_end

    if first_of_pay:
        return True, "DAY1", d
    if in_3_month_window:
        return True, "WINDOW_3M", d
    return False, "", d


# ─── Column Candidates ───────────────────────────────────────────────────────
YOUR_NID_CANDIDATES       = ["رقم قومي", "الرقم القومي", "رقم القومي", "قومي", "رقم قومى", "الرقم القومى"]
YOUR_MID_CANDIDATES       = ["رقم عسكري", "الرقم العسكري", "رقم العسكري", "عسكري", "رقم عسكرى", "الرقم العسكرى"]
YOUR_NAME_CANDIDATES      = ["اسم", "الاسم", "الاسم الكامل", "الأسماء", "أسماء", "اسم الفرد"]
YOUR_MODEL_CANDIDATES     = ["حالة النموذج", "نموذج", "حالة نموذج", "حاله النموذج"]
YOUR_STATUS_CANDIDATES    = ["حالة الفرد", "حالة فرد", "الفرد", "حاله الفرد"]
YOUR_DISCHARGE_CANDIDATES = ["تاريخ التسريح", "التسريح", "تسريح", "تاريخ تسريح"]
YOUR_NOTES_CANDIDATES     = ["ملاحظة", "ملاحظات", "ملاحظه"]

# ── NEW: Sector & Department ──────────────────────────────────────────────────
SECTOR_CANDIDATES         = ["القطاع", "قطاع", "اسم القطاع"]
DEPARTMENT_CANDIDATES     = ["الإدارة", "ادارة", "اسم الإدارة", "الادارة", "إدارة"]

THEIR_NID_CANDIDATES      = YOUR_NID_CANDIDATES
THEIR_MID_CANDIDATES      = YOUR_MID_CANDIDATES
THEIR_NAME_CANDIDATES     = YOUR_NAME_CANDIDATES
THEIR_AMOUNT_CANDIDATES   = ["المبلغ", "قيمة الحافز", "الحافز", "مبلغ الحافز", "المبلغ المستحق", "حوافز", "حافز"]


# ─── Worker Thread ────────────────────────────────────────────────────────────
class CompareWorker(QThread):
    progress = pyqtSignal(int, str)
    finished = pyqtSignal(str)
    error    = pyqtSignal(str)

    def __init__(self, your_file, their_file, sector_name, output_dir, payment_month=""):
        super().__init__()
        self.your_file      = your_file
        self.their_file     = their_file
        self.sector_name    = sector_name    # user-typed fallback
        self.output_dir     = output_dir
        self.payment_month  = payment_month  # format: "YYYY-MM"
        self.auto_sector    = ""             # filled during _compare()

    def run(self):
        try:
            self.progress.emit(10, "جاري قراءة الشيت الخاص بك…")
            your_df = self._read_smart(self.your_file)

            self.progress.emit(30, "جاري قراءة الشيت الخاص بهم…")
            their_df = self._read_smart(self.their_file)

            self.progress.emit(50, "جاري المطابقة والمقارنة…")
            result_df = self._compare(your_df, their_df)

            self.progress.emit(75, "جاري إنشاء ملف Excel…")
            output_path = self._export(result_df)

            self.progress.emit(100, "✅  تمت العملية بنجاح!")
            self.finished.emit(output_path)
        except Exception as e:
            import traceback
            self.error.emit(str(e) + "\n\n" + traceback.format_exc())

    def _read_smart(self, path):
        try:
            xl    = pd.ExcelFile(path)
            sheet = xl.sheet_names[0]
            for skip in range(20):
                df = pd.read_excel(path, sheet_name=sheet, skiprows=skip, dtype=str)
                df.columns = [str(c).strip() for c in df.columns]
                arabic = sum(1 for c in df.columns if any('\u0600' <= ch <= '\u06FF' for ch in c))
                if arabic >= 2:
                    return df.dropna(how='all').reset_index(drop=True)
            return pd.read_excel(path, sheet_name=sheet, dtype=str)
        except PermissionError:
            temp_dir  = tempfile.gettempdir()
            temp_file = os.path.join(temp_dir, os.path.basename(path))
            shutil.copy2(path, temp_file)
            return pd.read_excel(temp_file, dtype=str)
        except Exception as e:
            raise Exception(f"خطأ في قراءة الملف {path}: {str(e)}")

    # ──────────────────────────────────────────────────────────────────────────
    def _compare(self, your_df, their_df):
        # ── Detect columns in YOUR sheet ─────────────────────────────────────
        your_nid    = find_col(your_df, YOUR_NID_CANDIDATES)
        your_mid    = find_col(your_df, YOUR_MID_CANDIDATES)
        your_name   = find_col(your_df, YOUR_NAME_CANDIDATES)
        your_model  = find_col(your_df, YOUR_MODEL_CANDIDATES)
        your_status = find_col(your_df, YOUR_STATUS_CANDIDATES)
        your_disc   = find_col(your_df, YOUR_DISCHARGE_CANDIDATES)
        your_notes  = find_col(your_df, YOUR_NOTES_CANDIDATES)
        your_sector = find_col(your_df, SECTOR_CANDIDATES)        # ← NEW
        your_dept   = find_col(your_df, DEPARTMENT_CANDIDATES)    # ← NEW

        # Auto-detect sector name for report title (first non-empty value)
        if your_sector:
            first_vals = [
                str(v).strip()
                for v in your_df[your_sector].dropna()
                if str(v).strip() not in ("", "nan")
            ]
            if first_vals:
                self.auto_sector = first_vals[0]

        # ── Detect columns in THEIR sheet ────────────────────────────────────
        their_nid    = find_col(their_df, THEIR_NID_CANDIDATES)
        their_mid    = find_col(their_df, THEIR_MID_CANDIDATES)
        their_name   = find_col(their_df, THEIR_NAME_CANDIDATES)
        their_amount = find_col(their_df, THEIR_AMOUNT_CANDIDATES)

        if not their_mid and not their_nid:
            raise ValueError(
                "لم يُعثر على عمود الرقم العسكري أو الرقم القومي في الشيت الخاص بهم.\n"
                f"الأعمدة الموجودة: {list(their_df.columns)}"
            )

        # ── Build lookup dicts for THEIR sheet ───────────────────────────────
        their_by_mid = {}
        their_by_nid = {}
        for _, row in their_df.iterrows():
            if their_mid:
                mid_key = normalize_id(row.get(their_mid, ""))
                if mid_key:
                    their_by_mid[mid_key] = row
            if their_nid:
                nid_key = normalize_id(row.get(their_nid, ""))
                if nid_key:
                    their_by_nid[nid_key] = row

        # ── Build result rows ─────────────────────────────────────────────────
        rows    = []
        counter = 1
        for _, yrow in your_df.iterrows():
            r = {}
            r["seq"] = counter
            counter += 1

            your_nid_val  = yrow.get(your_nid,  "") if your_nid  else ""
            your_mid_val  = yrow.get(your_mid,  "") if your_mid  else ""
            your_name_val = yrow.get(your_name, "") if your_name else ""

            r["your_name"]    = your_name_val
            r["your_notes"]   = yrow.get(your_notes,  "") if your_notes  else ""
            r["your_model"]   = yrow.get(your_model,  "") if your_model  else ""
            r["your_status"]  = yrow.get(your_status, "") if your_status else ""
            # r["your_model"]   = yrow.get(your_model, "") if your_model else ""
            r["your_nid_raw"] = your_nid_val
            r["your_mid_raw"] = your_mid_val

            # sector & department per row
            sec_val = str(yrow.get(your_sector, "")).strip() if your_sector else ""
            dept_val = str(yrow.get(your_dept,   "")).strip() if your_dept   else ""
            if sec_val.lower() == "nan":
                sec_val = ""
            if dept_val.lower() == "nan":
                dept_val = ""
            r["your_sector"] = sec_val
            r["your_dept"]   = dept_val

            disc_raw       = yrow.get(your_disc, "") if your_disc else ""
            r["your_disc"] = format_date_only(disc_raw)

            not_el, disc_reason, _parsed = _discharge_ineligibility(
                str(disc_raw) if disc_raw is not None else "", self.payment_month
            )
            r["not_eligible"] = not_el
            r["disc_reason"]  = disc_reason  # DAY1 | WINDOW_3M | HARD_DATE | ""

            status_kind, status_label = classify_person_status(r["your_status"])
            r["status_kind"]  = status_kind
            r["status_label"] = status_label

            reviewed_sector = (self.auto_sector or self.sector_name or "").strip()
            if not sec_val:
                r["sector_ok"] = False
                r["sector_hard"] = True
            elif reviewed_sector:
                eq = names_equivalent(sec_val, reviewed_sector)
                r["sector_ok"] = bool(eq)
                r["sector_hard"] = eq is None
            else:
                r["sector_ok"] = True
                r["sector_hard"] = False

            # الإدارة must exist; without a master list of depts-per-sector,
            # empty / placeholder values are denied, odd values are hard cases.
            dept_norm = normalize_arabic(dept_val)
            placeholder = dept_norm in ("", "لاشيء", "بدون", "غيرمحدد", "-")
            r["dept_ok"] = bool(dept_norm) and not placeholder
            r["dept_hard"] = False
            if r["dept_ok"] and reviewed_sector and names_equivalent(dept_val, reviewed_sector):
                # department cell repeats the sector name → not a real إدارة
                r["dept_ok"] = False
                r["dept_hard"] = True

            # ── Match against THEIR sheet ────────────────────────────────────
            mid_key = normalize_id(your_mid_val)
            nid_key = normalize_id(your_nid_val)

            their_row = their_by_mid.get(mid_key) if mid_key else None
            if their_row is None and nid_key:
                their_row = their_by_nid.get(nid_key)

            if their_row is not None:
                r["found"] = True
                their_name_val   = their_row.get(their_name,   "") if their_name   else ""
                their_nid_val    = their_row.get(their_nid,    "") if their_nid    else ""
                their_mid_val    = their_row.get(their_mid,    "") if their_mid    else ""
                their_amount_val = their_row.get(their_amount, "") if their_amount else ""

                r["their_name"]   = their_name_val
                r["their_nid"]    = their_nid_val
                r["their_mid"]    = their_mid_val
                r["their_amount"] = their_amount_val

                nid_match = (nid_key == normalize_id(their_nid_val)) if nid_key else False
                mid_match = (mid_key == normalize_id(their_mid_val)) if mid_key else False
                r["nid_match"] = nid_match
                r["mid_match"] = mid_match
                r["correct_nid"] = your_nid_val if not nid_match else ""
                r["correct_mid"] = your_mid_val if not mid_match else ""

                nm = names_equivalent(your_name_val, their_name_val)
                r["name_match"] = bool(nm)
                r["name_hard"]  = nm is None
            else:
                r["found"]        = False
                r["their_name"]   = "غير موجود"
                r["their_nid"]    = ""
                r["their_mid"]    = ""
                r["their_amount"] = ""
                r["nid_match"]    = False
                r["mid_match"]    = False
                r["correct_nid"]  = your_nid_val
                r["correct_mid"]  = your_mid_val
                r["name_match"]   = False
                r["name_hard"]    = False

            # ── Final decision ───────────────────────────────────────────────
            # grant | deny | hard
            reasons = []
            decision = "grant"

            if not r["found"]:
                decision = "deny"
                reasons.append("غير موجود في شيت الصرف")
            else:
                if not r["nid_match"]:
                    decision = "hard"
                    reasons.append("اختلاف الرقم القومي")
                if not r["mid_match"]:
                    decision = "hard"
                    reasons.append("اختلاف الرقم العسكري")
                if r.get("name_hard"):
                    decision = "hard"
                    reasons.append("الاسم ناقص للمقارنة")
                elif r["found"] and r.get("name_match") is False:
                    # IDs may match but names do not → manual review
                    if decision != "deny":
                        decision = "hard"
                    reasons.append("الاسم غير مطابق مع الرقم")

            if r["status_kind"] == "deny":
                decision = "deny"
                reasons.append("حالة الفرد: خارج")
            elif r["status_kind"] == "hard":
                if decision != "deny":
                    decision = "hard"
                reasons.append("حالة الفرد تحتاج مراجعة")

            if r["not_eligible"]:
                decision = "deny"
                if r["disc_reason"] == "DAY1":
                    reasons.append("تسريح يوم 1 من شهر الصرف")
                else:
                    reasons.append("تسريح خلال 3 أشهر من شهر الصرف")
            elif r["disc_reason"] == "HARD_DATE":
                if decision != "deny":
                    decision = "hard"
                reasons.append("تاريخ التسريح غير مقروء")

            if not r["sector_ok"]:
                if r.get("sector_hard"):
                    if decision != "deny":
                        decision = "hard"
                    reasons.append("القطاع غير مذكور")
                else:
                    decision = "deny"
                    reasons.append("القطاع لا يطابق القطاع قيد المراجعة")

            if not r["dept_ok"]:
                if r.get("dept_hard"):
                    if decision != "deny":
                        decision = "hard"
                    reasons.append("الإدارة ليست إدارة حقيقية داخل القطاع")
                else:
                    decision = "deny"
                    reasons.append("الإدارة غير محددة")

            r["decision"] = decision
            r["decision_reasons"] = " + ".join(reasons) if reasons else "مطابق ومستحق"

            if decision == "grant":
                r["status_text"] = "✅ مطابق ومستحق"
            elif decision == "hard":
                r["status_text"] = "🔍 يحتاج مراجعة: " + r["decision_reasons"]
            else:
                r["status_text"] = "❌ غير مستحق: " + r["decision_reasons"]

            rows.append(r)

        result_df = pd.DataFrame(rows)
        # Keep EVERY row (including hard cases) so they can be handled manually.
        if not result_df.empty:
            result_df["seq"] = range(1, len(result_df) + 1)
        return result_df

    # ──────────────────────────────────────────────────────────────────────────
    def _export(self, df):
        today    = datetime.now().strftime("%Y-%m-%d")
        safe_sec = self.sector_name.strip().replace("/", "-").replace("\\", "-")
        filename = f"مراجعة حافز {safe_sec} {today}.xlsx"

        if not ensure_directory_permission(self.output_dir):
            self.output_dir = get_safe_output_directory()

        out_path = os.path.join(self.output_dir, filename)

        wb = Workbook()
        ws = wb.active
        ws.title = "مراجعة الحافز"
        ws.sheet_view.rightToLeft = True

        # ── Colours ──────────────────────────────────────────────────────────
        DARK_NAVY        = "1B2A4A"
        LIGHT_GOLD       = "F0D080"
        YOUR_BG          = "1F4E79"
        THEIR_BG         = "2E7D32"
        REVIEW_BG        = "8D6E63"
        CORRECT_BG       = "FF6F00"
        MATCH_GRN        = "E8F5E9"
        MISMATCH_ORG     = "FFF3E0"
        NOTFOUND_RED     = "FFEBEE"
        NOT_ELIGIBLE_BG  = "EDE7F6"
        NOT_ELIGIBLE_FG  = "4A148C"
        NOT_ELIGIBLE_HDR = "6A1B9A"
        HARD_BG          = "FFF8E1"
        HARD_FG          = "E65100"
        HARD_HDR         = "F9A825"
        HEADER_BG        = "37474F"
        BORDER_CLR       = "B0BEC5"

        thin  = Side(style='thin',   color=BORDER_CLR)
        thick = Side(style='medium', color=DARK_NAVY)
        brd   = Border(left=thin,  right=thin,  top=thin,  bottom=thin)
        brd_h = Border(left=thick, right=thick, top=thick, bottom=thick)

        def cs(cell, bg=None, fg="000000", bold=False, sz=10,
               ha="center", b=brd, wrap=True):
            if bg:
                cell.fill = PatternFill("solid", fgColor=bg)
            cell.font      = Font(name="Arial", bold=bold, color=fg, size=sz)
            cell.alignment = Alignment(horizontal=ha, vertical="center",
                                       wrap_text=wrap)
            cell.border    = b

        def add_comment(cell, text, author="النظام"):
            c = Comment(text, author)
            c.width  = 260
            c.height = 85
            cell.comment = c

        TOTAL_COLS = 16

        # Prefer auto-detected sector from data; fall back to user-typed value
        report_sector = self.auto_sector if self.auto_sector else self.sector_name

        # ── Row 1: Main title ─────────────────────────────────────────────────
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=TOTAL_COLS)
        cs(ws.cell(1, 1, f"تقرير مراجعة حوافز قطاع: {report_sector}"),
           DARK_NAVY, LIGHT_GOLD, True, 16, b=brd_h)

        # ── Row 2: Group headers ──────────────────────────────────────────────
        cs(ws.cell(2, 1, "م"), HEADER_BG, "FFFFFF", True, 10, b=brd_h)

        ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=6)
        cs(ws.cell(2, 2, "بيانات الشيت الخاص بك"),
           YOUR_BG, LIGHT_GOLD, True, 12, b=brd_h)

        ws.merge_cells(start_row=2, start_column=7, end_row=2, end_column=9)
        cs(ws.cell(2, 7, "بيانات المراجعة (من شيتك)"),
           REVIEW_BG, LIGHT_GOLD, True, 12, b=brd_h)

        ws.merge_cells(start_row=2, start_column=10, end_row=2, end_column=13)
        cs(ws.cell(2, 10, "بيانات الشيت الخاص بهم"),
           THEIR_BG, LIGHT_GOLD, True, 12, b=brd_h)

        ws.merge_cells(start_row=2, start_column=14, end_row=2, end_column=15)
        cs(ws.cell(2, 14, "بيانات التصحيح المقترحة"),
           CORRECT_BG, "FFFFFF", True, 12, b=brd_h)

        cs(ws.cell(2, 16, "قرار الاستحقاق"), HEADER_BG, "FFFFFF", True, 12, b=brd_h)

        # ── Row 3: Column labels ──────────────────────────────────────────────
        headers_row3 = [
            ("م",                       HEADER_BG),
            ("الاسم\n(بالرقم القومي)",  YOUR_BG),
            ("الاسم\n(بالرقم العسكري)", YOUR_BG),
            ("القطاع",                  YOUR_BG),
            ("الإدارة",                 YOUR_BG),
            ("ملاحظة",                  YOUR_BG),
            ("حالة النموذج",             REVIEW_BG),
            ("حالة الفرد",               REVIEW_BG),
            ("تاريخ التسريح",            REVIEW_BG),
            ("الاسم",                   THEIR_BG),
            ("الرقم القومي",             THEIR_BG),
            ("الرقم العسكري",            THEIR_BG),
            ("المبلغ",                  THEIR_BG),
            ("تصحيح الرقم القومي",       CORRECT_BG),
            ("تصحيح الرقم العسكري",      CORRECT_BG),
            ("قرار الاستحقاق",           HEADER_BG),
        ]
        for ci, (header, bg) in enumerate(headers_row3, start=1):
            cs(ws.cell(3, ci, header), bg, "FFFFFF", True, 9, b=brd_h)

        ws.row_dimensions[1].height = 32
        ws.row_dimensions[2].height = 30
        ws.row_dimensions[3].height = 44

        # ── Data rows ─────────────────────────────────────────────────────────
        for ri, (_, row) in enumerate(df.iterrows(), start=4):
            found        = bool(row.get("found",        False))
            mid_match    = bool(row.get("mid_match",    False))
            nid_match    = bool(row.get("nid_match",    False))
            name_match   = bool(row.get("name_match",   False))
            not_eligible = bool(row.get("not_eligible", False))
            decision     = str(row.get("decision", "hard"))

            if decision == "deny" and not_eligible:
                row_bg = NOT_ELIGIBLE_BG
            elif decision == "deny":
                row_bg = NOTFOUND_RED
            elif decision == "hard":
                row_bg = HARD_BG
            else:
                row_bg = MATCH_GRN if (ri % 2 == 0) else "FFFFFF"

            def v(key):
                val = row.get(key, "")
                return str(val) if pd.notna(val) else ""

            cs(ws.cell(ri, 1, v("seq")), row_bg, sz=9)

            if not found:
                name_nid_bg, name_nid_fg, name_nid_bold = NOTFOUND_RED, "C62828", True
            elif not nid_match or not name_match:
                name_nid_bg, name_nid_fg, name_nid_bold = MISMATCH_ORG, "E65100", True
            elif decision == "deny":
                name_nid_bg, name_nid_fg, name_nid_bold = row_bg, NOT_ELIGIBLE_FG if not_eligible else "C62828", False
            else:
                name_nid_bg, name_nid_fg, name_nid_bold = MATCH_GRN, "1B5E20", False

            nid_cell = ws.cell(ri, 2, v("your_name"))
            cs(nid_cell, name_nid_bg, name_nid_fg, bold=name_nid_bold, sz=9)
            if not found:
                add_comment(nid_cell, "غير موجود في شيت الصرف")
            elif not nid_match or not name_match:
                add_comment(
                    nid_cell,
                    f"الاسم عندك: {v('your_name')}\n"
                    f"الاسم عندهم: {v('their_name')}\n"
                    f"رقمك القومي:  {v('your_nid_raw')}\n"
                    f"رقمهم القومي: {v('their_nid')}"
                )

            if not found:
                name_mid_bg, name_mid_fg, name_mid_bold = NOTFOUND_RED, "C62828", True
            elif not mid_match or not name_match:
                name_mid_bg, name_mid_fg, name_mid_bold = MISMATCH_ORG, "E65100", True
            elif decision == "deny":
                name_mid_bg, name_mid_fg, name_mid_bold = row_bg, NOT_ELIGIBLE_FG if not_eligible else "C62828", False
            else:
                name_mid_bg, name_mid_fg, name_mid_bold = MATCH_GRN, "1B5E20", False

            mid_cell = ws.cell(ri, 3, v("your_name"))
            cs(mid_cell, name_mid_bg, name_mid_fg, bold=name_mid_bold, sz=9)
            if not found:
                add_comment(mid_cell, "غير موجود في شيت الصرف")
            elif not mid_match or not name_match:
                add_comment(
                    mid_cell,
                    f"الاسم عندك: {v('your_name')}\n"
                    f"الاسم عندهم: {v('their_name')}\n"
                    f"رقمك العسكري:  {v('your_mid_raw')}\n"
                    f"رقمهم العسكري: {v('their_mid')}"
                )

            sec_cell = ws.cell(ri, 4, v("your_sector"))
            if not bool(row.get("sector_ok", True)):
                cs(sec_cell, MISMATCH_ORG, "E65100", True, sz=9)
                add_comment(sec_cell, "القطاع لا يطابق القطاع قيد المراجعة")
            else:
                cs(sec_cell, row_bg, sz=9)

            dept_cell = ws.cell(ri, 5, v("your_dept"))
            if not bool(row.get("dept_ok", True)):
                cs(dept_cell, MISMATCH_ORG, "E65100", True, sz=9)
                add_comment(dept_cell, "الإدارة فارغة أو لا تبدو إدارة داخل القطاع")
            else:
                cs(dept_cell, row_bg, sz=9)

            cs(ws.cell(ri, 6, v("your_notes")), row_bg, sz=9)
            cs(ws.cell(ri, 7, v("your_model")),  row_bg, sz=9)

            st_cell = ws.cell(ri, 8, v("your_status"))
            sk = v("status_kind")
            if sk == "deny":
                cs(st_cell, NOTFOUND_RED, "C62828", True, sz=9)
                add_comment(st_cell, "خارج → غير مستحق (إلا مجلس طبي)")
            elif sk == "hard":
                cs(st_cell, HARD_BG, HARD_FG, True, sz=9)
            else:
                cs(st_cell, row_bg, sz=9)

            disc_cell = ws.cell(ri, 9, v("your_disc"))
            if not_eligible:
                cs(disc_cell, NOT_ELIGIBLE_BG, NOT_ELIGIBLE_FG, bold=True, sz=9)
                why = v("disc_reason")
                if why == "DAY1":
                    add_comment(disc_cell, "تسريح يوم 1 من شهر الصرف → لا يستحق")
                else:
                    add_comment(disc_cell, "التسريح داخل مدة 3 أشهر من شهر الصرف → لا يستحق")
            elif v("disc_reason") == "HARD_DATE":
                cs(disc_cell, HARD_BG, HARD_FG, bold=True, sz=9)
                add_comment(disc_cell, "تعذر قراءة تاريخ التسريح — للمراجعة اليدوية")
            else:
                cs(disc_cell, row_bg, sz=9)

            cs(ws.cell(ri, 10, v("their_name")), row_bg, sz=9)

            their_nid_cell = ws.cell(ri, 11, v("their_nid"))
            if not nid_match and found:
                cs(their_nid_cell, MISMATCH_ORG, "E65100", bold=True, sz=9)
            else:
                cs(their_nid_cell, row_bg, sz=9)

            their_mid_cell = ws.cell(ri, 12, v("their_mid"))
            if not mid_match and found:
                cs(their_mid_cell, MISMATCH_ORG, "E65100", bold=True, sz=9)
            else:
                cs(their_mid_cell, row_bg, sz=9)

            cs(ws.cell(ri, 13, v("their_amount")), row_bg, "DAA520", True, sz=9)

            correct_nid = v("correct_nid")
            if correct_nid and not nid_match:
                cs(ws.cell(ri, 14, correct_nid), MISMATCH_ORG, "E65100", True, sz=9)
            else:
                cs(ws.cell(ri, 14, correct_nid), row_bg, sz=9)

            correct_mid = v("correct_mid")
            if correct_mid and not mid_match:
                cs(ws.cell(ri, 15, correct_mid), MISMATCH_ORG, "E65100", True, sz=9)
            else:
                cs(ws.cell(ri, 15, correct_mid), row_bg, sz=9)

            status_txt = v("status_text") or v("decision_reasons")
            if decision == "grant":
                cs(ws.cell(ri, 16, status_txt), row_bg, "1B5E20", True, sz=8)
            elif decision == "hard":
                cs(ws.cell(ri, 16, status_txt), HARD_BG, HARD_FG, True, sz=8)
            elif not_eligible:
                cs(ws.cell(ri, 16, status_txt), NOT_ELIGIBLE_BG, NOT_ELIGIBLE_FG, True, sz=8)
            else:
                cs(ws.cell(ri, 16, status_txt), NOTFOUND_RED, "C62828", True, sz=8)

            ws.row_dimensions[ri].height = 28

        # ── Summary row ───────────────────────────────────────────────────────
        lr     = 4 + len(df)
        total  = len(df)
        n_grant = sum(1 for _, r in df.iterrows() if str(r.get("decision", "")) == "grant")
        n_hard  = sum(1 for _, r in df.iterrows() if str(r.get("decision", "")) == "hard")
        n_deny  = sum(1 for _, r in df.iterrows() if str(r.get("decision", "")) == "deny")
        n_disc  = sum(1 for _, r in df.iterrows() if bool(r.get("not_eligible", False)))

        ws.merge_cells(start_row=lr, start_column=1,  end_row=lr, end_column=3)
        cs(ws.cell(lr, 1, f"📊 الإجمالي: {total}"),
           DARK_NAVY, LIGHT_GOLD, True, 11, b=brd_h)

        ws.merge_cells(start_row=lr, start_column=4,  end_row=lr, end_column=6)
        cs(ws.cell(lr, 4, f"✅ مستحق: {n_grant}"),
           THEIR_BG, "FFFFFF", True, 11, b=brd_h)

        ws.merge_cells(start_row=lr, start_column=7,  end_row=lr, end_column=9)
        cs(ws.cell(lr, 7, f"🔍 للمراجعة اليدوية: {n_hard}"),
           HARD_HDR, "1B2A4A", True, 11, b=brd_h)

        ws.merge_cells(start_row=lr, start_column=10, end_row=lr, end_column=12)
        cs(ws.cell(lr, 10, f"❌ غير مستحق: {n_deny}"),
           "C62828", "FFFFFF", True, 11, b=brd_h)

        ws.merge_cells(start_row=lr, start_column=13, end_row=lr, end_column=TOTAL_COLS)
        cs(ws.cell(lr, 13, f"🚫 منها تسريح (يوم 1 / 3 أشهر): {n_disc}"),
           NOT_ELIGIBLE_HDR, "FFFFFF", True, 11, b=brd_h)

        ws.row_dimensions[lr].height = 30

        # ── Column widths ─────────────────────────────────────────────────────
        col_widths = {
            1:  6, 2:  28, 3:  28, 4:  18, 5:  18, 6:  20, 7:  18, 8:  18,
            9:  20, 10: 26, 11: 20, 12: 20, 13: 14, 14: 22, 15: 22, 16: 42,
        }
        for col_idx, width in col_widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.freeze_panes = "A4"

        # ── Save ──────────────────────────────────────────────────────────────
        def try_save(path):
            wb.save(path)

        saved = False
        last_err = None
        for attempt in range(3):
            try:
                try_save(out_path)
                saved = True
                break
            except PermissionError as e:
                last_err = e
                # File is likely open elsewhere / locked. Try a fresh, unique
                # filename instead of overwriting, so we don't get stuck
                # retrying the same locked path.
                ts = datetime.now().strftime("%H%M%S")
                alt_filename = f"مراجعة حافز {safe_sec} {today} ({ts}).xlsx"
                out_path = os.path.join(self.output_dir, alt_filename)
            except Exception as e:
                last_err = e
                break

        if not saved:
            # Last resort: save into the OS temp folder, which is always writable.
            try:
                alt_filename = f"مراجعة حافز {safe_sec} {today} ({datetime.now().strftime('%H%M%S')}).xlsx"
                fallback_path = os.path.join(tempfile.gettempdir(), alt_filename)
                wb.save(fallback_path)
                out_path = fallback_path
                saved = True
            except Exception as e:
                raise Exception(
                    "تعذّر حفظ الملف بسبب مشكلة صلاحيات/قفل على الملف.\n"
                    "تأكد من إغلاق أي نسخة مفتوحة من ملف الإخراج فى Excel ثم أعد المحاولة.\n\n"
                    f"تفاصيل الخطأ: {last_err or e}"
                )

        return out_path


# ─── Drop Zone ───────────────────────────────────────────────────────────────
class DropZone(QFrame):
    def __init__(self, label, night=False, parent=None):
        super().__init__(parent)
        self.file_path  = ""
        self.label_text = label
        self.night      = night
        self.setAcceptDrops(True)
        self.setMinimumHeight(120)
        self.setCursor(Qt.CursorShape.PointingHandCursor)
        self._build()
        self._style_empty()

    def set_night(self, night):
        self.night = night
        if self.file_path:
            self._style_loaded()
        else:
            self._style_empty()

    def _build(self):
        lay = QVBoxLayout(self)
        lay.setAlignment(Qt.AlignmentFlag.AlignCenter)

        self.ico = QLabel("📂")
        self.ico.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.ico.setFont(QFont("Segoe UI Emoji", 26))

        self.main = QLabel(self.label_text)
        self.main.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.main.setFont(QFont("Arial", 10, QFont.Weight.Bold))
        self.main.setWordWrap(True)

        self.sub = QLabel("اسحب الملف هنا أو انقر للاختيار")
        self.sub.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.sub.setFont(QFont("Arial", 8))

        self.fname = QLabel("")
        self.fname.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.fname.setFont(QFont("Arial", 8))
        self.fname.setWordWrap(True)

        for w in [self.ico, self.main, self.sub, self.fname]:
            lay.addWidget(w)

    def _style_empty(self):
        if self.night:
            self.setStyleSheet("""
                DropZone { border:2px dashed #4A6FA5; border-radius:12px; background:#1E2A3A; }
                DropZone:hover { border-color:#7EB8F7; background:#1A2535; }
                QLabel { color:#90CAF9; background:transparent; border:none; }
            """)
        else:
            self.setStyleSheet("""
                DropZone { border:2px dashed #90CAF9; border-radius:12px; background:#F0F7FF; }
                DropZone:hover { border-color:#1B2A4A; background:#E3F0FF; }
                QLabel { color:#1B2A4A; background:transparent; border:none; }
            """)
        self.ico.setText("📂")
        self.sub.setVisible(True)

    def _style_loaded(self):
        if self.night:
            self.setStyleSheet("""
                DropZone { border:2px solid #4CAF50; border-radius:12px; background:#1A2E1A; }
                QLabel { color:#A5D6A7; background:transparent; border:none; }
            """)
        else:
            self.setStyleSheet("""
                DropZone { border:2px solid #2E7D32; border-radius:12px; background:#E8F5E9; }
                QLabel { color:#1B5E20; background:transparent; border:none; }
            """)
        self.ico.setText("✅")
        self.sub.setVisible(False)

    def mousePressEvent(self, e):
        self._pick()

    def dragEnterEvent(self, e):
        if e.mimeData().hasUrls():
            e.acceptProposedAction()

    def dropEvent(self, e):
        urls = e.mimeData().urls()
        if urls:
            p = urls[0].toLocalFile()
            if p.lower().endswith(('.xlsx', '.xls')):
                self._load(p)
            else:
                QMessageBox.warning(self, "خطأ", "يرجى اختيار ملف Excel (.xlsx أو .xls)")

    def _pick(self):
        p, _ = QFileDialog.getOpenFileName(
            self, f"اختر {self.label_text}", "", "Excel Files (*.xlsx *.xls)"
        )
        if p:
            self._load(p)

    def _load(self, path):
        self.file_path = path
        self.fname.setText(f"📄  {os.path.basename(path)}")
        self._style_loaded()


# ─── Main Window ──────────────────────────────────────────────────────────────
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.night_mode = False
        self.setWindowTitle("منظومة الحوافز - نظام المطابقة والتصحيح")
        self.setMinimumSize(1400, 700)
        self.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
        self._build_ui()
        self._apply_theme()
        self._load_logo()

        enter_sc  = QShortcut(QKeySequence(Qt.Key.Key_Return), self)
        enter_sc.activated.connect(self._run)
        enter_sc2 = QShortcut(QKeySequence(Qt.Key.Key_Enter), self)
        enter_sc2.activated.connect(self._run)

    def _load_logo(self):
        try:
            possible_paths = [
                os.path.join(os.path.dirname(os.path.abspath(__file__)),
                             "جهاز_مستقبل_مصر_للتنمية_المستدامة.png"),
                os.path.join(os.path.expanduser("~/Desktop"),
                             "جهاز_مستقبل_مصر_للتنمية_المستدامة.png"),
                "جهاز_مستقبل_مصر_للتنمية_المستدامة.png"
            ]
            for logo_path in possible_paths:
                if os.path.exists(logo_path):
                    try:
                        pixmap = QPixmap(logo_path)
                        if not pixmap.isNull():
                            scaled = pixmap.scaled(
                                60, 60,
                                Qt.AspectRatioMode.KeepAspectRatio,
                                Qt.TransformationMode.SmoothTransformation
                            )
                            self.logo_label.setPixmap(scaled)
                            self.logo_label.setVisible(True)
                            return
                    except Exception:
                        continue
            self.logo_label.setText("🏛️\nمصر")
            self.logo_label.setFont(QFont("Segoe UI Emoji", 24))
            self.logo_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
            self.logo_label.setStyleSheet("""
                QLabel { color:#C9A84C;
                         background:qlineargradient(x1:0,y1:0,x2:1,y2:1,
                             stop:0 #1B2A4A,stop:1 #2A5090);
                         border-radius:30px; font-weight:bold; }
            """)
            self.logo_label.setVisible(True)
        except Exception:
            self.logo_label.setVisible(False)

    def _apply_theme(self):
        n = self.night_mode
        BG       = "#0F1923" if n else "#F5F7FA"
        CARD     = "#1A2535" if n else "#FFFFFF"
        BORDER   = "#2A3A50" if n else "#E0E0E0"
        TXT      = "#D0E4F7" if n else "#1B2A4A"
        INP_BG   = "#111C28" if n else "#FFFFFF"
        INP_BDR  = "#3A5070" if n else "#B0BEC5"
        INP_FOC  = "#5A8FCC" if n else "#1B2A4A"
        PB_BG    = "#1A2535" if n else "#E0E0E0"
        TXT_SUB  = "#7EB8F7" if n else "#546E7A"
        NIGHT_BTN = ("#2A3A50", "#D0E4F7") if n else ("#E3F0FF", "#1B2A4A")

        self.setStyleSheet(f"""
            QMainWindow {{ background:{BG}; }}
            QWidget#central {{ background:{BG}; }}
            QScrollArea {{ border:none; background:{BG}; }}
            QWidget#inner {{ background:{BG}; }}
            QFrame#card {{
                background:{CARD}; border-radius:14px; border:1px solid {BORDER};
            }}
            QLabel#secTitle {{
                color:{TXT}; font-size:12px; font-weight:bold; font-family:Arial;
            }}
            QLabel#subLbl {{
                color:{TXT_SUB}; font-family:Arial; font-size:10px;
            }}
            QLineEdit {{
                border:2px solid {INP_BDR}; border-radius:8px; padding:8px 12px;
                font-size:13px; font-family:Arial; background:{INP_BG}; color:{TXT};
            }}
            QLineEdit:focus {{ border-color:{INP_FOC}; }}
            QPushButton#runBtn {{
                background:qlineargradient(x1:0,y1:0,x2:1,y2:0,
                    stop:0 #1B2A4A, stop:1 #2A5090);
                color:#F0D080; border:none; border-radius:12px;
                padding:18px 40px; font-size:17px; font-weight:bold; font-family:Arial;
            }}
            QPushButton#runBtn:hover  {{ background:#2A5090; }}
            QPushButton#runBtn:pressed {{ background:#0F1E38; }}
            QPushButton#runBtn:disabled {{ background:#555; color:#888; }}
            QPushButton#outBtn {{
                background:transparent; color:{TXT}; border:2px solid {INP_BDR};
                border-radius:8px; padding:8px 18px; font-size:11px; font-family:Arial;
            }}
            QPushButton#outBtn:hover {{
                background:{NIGHT_BTN[0]}; color:{NIGHT_BTN[1]};
            }}
            QPushButton#nightBtn {{
                background:{NIGHT_BTN[0]}; color:{NIGHT_BTN[1]};
                border:2px solid {INP_BDR}; border-radius:8px;
                padding:6px 14px; font-size:11px; font-family:Arial;
            }}
            QPushButton#nightBtn:hover {{ border-color:{INP_FOC}; }}
            QProgressBar {{
                border:none; border-radius:8px; background:{PB_BG}; height:18px;
                text-align:center; font-family:Arial; font-size:10px; color:white;
            }}
            QProgressBar::chunk {{
                background:qlineargradient(x1:0,y1:0,x2:1,y2:0,
                    stop:0 #1B2A4A, stop:1 #C9A84C);
                border-radius:8px;
            }}
        """)

        for dz in [self.your_drop, self.their_drop]:
            dz.set_night(self.night_mode)

        self.night_btn.setText("☀️  وضع النهار" if self.night_mode else "🌙  وضع الليل")

        header_grad = (
            "qlineargradient(x1:0,y1:0,x2:1,y2:0,"
            "stop:0 #060E18,stop:0.5 #0F1923,stop:1 #060E18)"
            if n else
            "qlineargradient(x1:0,y1:0,x2:1,y2:0,"
            "stop:0 #0F1E38,stop:0.5 #1B2A4A,stop:1 #0F1E38)"
        )
        self.header.setStyleSheet(
            f"QFrame {{ background:{header_grad}; border-bottom:3px solid #C9A84C; }}"
        )
        for card in self.findChildren(QFrame, "card"):
            eff = card.graphicsEffect()
            if eff:
                eff.setColor(QColor(0, 0, 0, 60 if n else 30))

    def _card(self, widget, title=None):
        card   = QFrame()
        card.setObjectName("card")
        shadow = QGraphicsDropShadowEffect()
        shadow.setBlurRadius(18)
        shadow.setColor(QColor(0, 0, 0, 30))
        shadow.setOffset(0, 3)
        card.setGraphicsEffect(shadow)
        lay = QVBoxLayout(card)
        lay.setContentsMargins(20, 14, 20, 14)
        lay.setSpacing(10)
        if title:
            lbl = QLabel(title)
            lbl.setObjectName("secTitle")
            lay.addWidget(lbl)
        lay.addWidget(widget)
        return card

    def _build_ui(self):
        central = QWidget()
        central.setObjectName("central")
        self.setCentralWidget(central)
        root = QVBoxLayout(central)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        # ── Header ────────────────────────────────────────────────────────────
        self.header = QFrame()
        self.header.setFixedHeight(110)
        hlay = QHBoxLayout(self.header)
        hlay.setContentsMargins(24, 0, 24, 0)

        self.logo_label = QLabel()
        self.logo_label.setFixedSize(70, 70)
        self.logo_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.logo_label.setStyleSheet("background:transparent; border:none;")

        logo_l = QLabel("⭐")
        logo_l.setFont(QFont("Segoe UI Emoji", 28))
        logo_l.setStyleSheet("color:#C9A84C; background:transparent; border:none;")

        title_col = QVBoxLayout()
        title_col.setSpacing(2)
        t = QLabel("منظومة الحوافز")
        t.setFont(QFont("Arial", 20, QFont.Weight.Bold))
        t.setStyleSheet("color:#F0D080; background:transparent; border:none;")
        t.setAlignment(Qt.AlignmentFlag.AlignCenter)
        sub = QLabel("نظام المطابقة والتصحيح المتقدم")
        sub.setFont(QFont("Arial", 10))
        sub.setStyleSheet("color:#90CAF9; background:transparent; border:none;")
        sub.setAlignment(Qt.AlignmentFlag.AlignCenter)
        sub2 = QLabel("جهاز مستقبل مصر للتنمية المستدامة")
        sub2.setFont(QFont("Arial", 9))
        sub2.setStyleSheet(
            "color:#C9A84C; background:transparent; border:none; font-weight:bold;"
        )
        sub2.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title_col.addWidget(t)
        title_col.addWidget(sub)
        title_col.addWidget(sub2)

        logo_r = QLabel("⭐")
        logo_r.setFont(QFont("Segoe UI Emoji", 28))
        logo_r.setStyleSheet("color:#C9A84C; background:transparent; border:none;")
        logo_r.setAlignment(Qt.AlignmentFlag.AlignRight)

        self.night_btn = QPushButton("🌙  وضع الليل")
        self.night_btn.setObjectName("nightBtn")
        self.night_btn.setFixedWidth(140)
        self.night_btn.clicked.connect(self._toggle_night)

        hlay.addWidget(logo_r)
        hlay.addLayout(title_col)
        hlay.addWidget(logo_l)
        hlay.addWidget(self.logo_label)
        hlay.addSpacing(16)
        hlay.addWidget(self.night_btn)

        root.addWidget(self.header)

        # ── Scroll content ────────────────────────────────────────────────────
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        inner = QWidget()
        inner.setObjectName("inner")
        content = QVBoxLayout(inner)
        content.setContentsMargins(28, 20, 28, 20)
        content.setSpacing(14)
        scroll.setWidget(inner)
        root.addWidget(scroll)

        # ── Sector name + Payment month (shared card) ─────────────────────────
        sec_w = QWidget()
        sec_w.setStyleSheet("background:transparent; border:none;")
        sec_lay = QVBoxLayout(sec_w)
        sec_lay.setContentsMargins(0, 0, 0, 0)
        sec_lay.setSpacing(10)

        row1 = QWidget()
        row1.setStyleSheet("background:transparent; border:none;")
        r1l  = QHBoxLayout(row1)
        r1l.setContentsMargins(0, 0, 0, 0)
        r1l.setSpacing(10)
        lbl_sec = QLabel("اسم القطاع:")
        lbl_sec.setFont(QFont("Arial", 11, QFont.Weight.Bold))
        lbl_sec.setStyleSheet("color:#1B2A4A; background:transparent; border:none;")
        lbl_sec.setSizePolicy(QSizePolicy.Policy.Fixed, QSizePolicy.Policy.Preferred)
        self.sector_input = QLineEdit()
        self.sector_input.setPlaceholderText(
            "مثال: قطاع الشرق  —  (احتياطي: يُستخدم إن لم يوجد عمود القطاع بالشيت)"
        )
        self.sector_input.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
        r1l.addWidget(self.sector_input)
        r1l.addWidget(lbl_sec)
        sec_lay.addWidget(row1)

        row2 = QWidget()
        row2.setStyleSheet("background:transparent; border:none;")
        r2l  = QHBoxLayout(row2)
        r2l.setContentsMargins(0, 0, 0, 0)
        r2l.setSpacing(10)
        lbl_pm = QLabel("شهر الصرف:")
        lbl_pm.setFont(QFont("Arial", 11, QFont.Weight.Bold))
        lbl_pm.setStyleSheet("color:#1B2A4A; background:transparent; border:none;")
        lbl_pm.setSizePolicy(QSizePolicy.Policy.Fixed, QSizePolicy.Policy.Preferred)
        self.payment_month_input = QLineEdit()
        self.payment_month_input.setPlaceholderText(
            "YYYY-MM  —  مثال: 2026-09  (إن تُرك فارغاً يُستخدم الشهر الحالي)"
        )
        self.payment_month_input.setLayoutDirection(Qt.LayoutDirection.LeftToRight)
        self.payment_month_input.setMaximumWidth(320)
        r2l.addStretch()
        r2l.addWidget(self.payment_month_input)
        r2l.addWidget(lbl_pm)
        sec_lay.addWidget(row2)

        content.addWidget(self._card(sec_w, "✏️  بيانات القطاع وشهر الصرف"))

        # ── Drop zones ────────────────────────────────────────────────────────
        zones = QHBoxLayout()
        zones.setSpacing(14)
        self.your_drop  = DropZone(
            "الشيت الخاص بك\n"
            "(الاسم + الرقم القومي + الرقم العسكري + "
            "حالة النموذج + حالة الفرد + تاريخ التسريح + ملاحظة)"
        )
        self.their_drop = DropZone(
            "الشيت الخاص بهم\n"
            "(الاسم + الرقم القومي + الرقم العسكري + المبلغ)"
        )
        zones.addWidget(self._card(self.their_drop, "📄  الشيت الخاص بهم"))
        zones.addWidget(self._card(self.your_drop,  "📋  الشيت الخاص بك"))
        content.addLayout(zones)

        # ── Output folder ─────────────────────────────────────────────────────
        out_w = QWidget()
        out_w.setStyleSheet("background:transparent; border:none;")
        ol = QHBoxLayout(out_w)
        ol.setContentsMargins(0, 0, 0, 0)
        ol.setSpacing(10)
        self.out_path = QLineEdit()
        self.out_path.setReadOnly(True)
        self.out_path.setText(get_safe_output_directory())
        out_btn = QPushButton("تغيير")
        out_btn.setObjectName("outBtn")
        out_btn.setFixedWidth(80)
        out_btn.clicked.connect(self._pick_out)
        ol.addWidget(self.out_path)
        ol.addWidget(out_btn)
        content.addWidget(self._card(out_w, "📁  مجلد الحفظ"))

        # ── Progress ──────────────────────────────────────────────────────────
        pg_w = QWidget()
        pg_w.setStyleSheet("background:transparent; border:none;")
        pl = QVBoxLayout(pg_w)
        pl.setContentsMargins(0, 0, 0, 0)
        pl.setSpacing(8)
        self.progress = QProgressBar()
        self.progress.setValue(0)
        self.status_label = QLabel(
            "جاهز للبدء  •  اضغط Enter أو زر البدء للتشغيل"
        )
        self.status_label.setObjectName("subLbl")
        self.status_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        pl.addWidget(self.progress)
        pl.addWidget(self.status_label)
        content.addWidget(self._card(pg_w))

        # ── Run button ────────────────────────────────────────────────────────
        self.run_btn = QPushButton(
            "🔍     بدء المقارنة وإنشاء التقرير     ( Enter )"
        )
        self.run_btn.setObjectName("runBtn")
        self.run_btn.setMinimumHeight(64)
        self.run_btn.clicked.connect(self._run)
        content.addWidget(self.run_btn)

        # ── Footer ────────────────────────────────────────────────────────────
        footer = QLabel("تم التطوير بواسطة عسكرى / ابراهيم شفيق")
        footer.setAlignment(Qt.AlignmentFlag.AlignCenter)
        footer.setFont(QFont("Arial", 9))
        footer.setStyleSheet("color:#90A4AE; background:transparent; padding:6px;")
        content.addWidget(footer)
        content.addStretch()

    def _toggle_night(self):
        self.night_mode = not self.night_mode
        self._apply_theme()

    def _pick_out(self):
        p = QFileDialog.getExistingDirectory(
            self, "اختر مجلد الحفظ", self.out_path.text()
        )
        if p and ensure_directory_permission(p):
            self.out_path.setText(p)
        elif p:
            QMessageBox.warning(
                self, "تنبيه",
                "لا يمكن الكتابة في هذا المجلد، سيتم استخدام مجلد آخر"
            )
            self.out_path.setText(get_safe_output_directory())

    def _run(self):
        if not self.your_drop.file_path:
            QMessageBox.warning(self, "تنبيه", "يرجى اختيار الشيت الخاص بك أولاً")
            return
        if not self.their_drop.file_path:
            QMessageBox.warning(self, "تنبيه", "يرجى اختيار الشيت الخاص بهم أولاً")
            return
        sec = self.sector_input.text().strip()
        if not sec:
            QMessageBox.warning(self, "تنبيه", "يرجى إدخال اسم القطاع")
            return

        pm = self.payment_month_input.text().strip()
        if pm:
            import re as _re
            if not _re.match(r'^\d{4}-\d{2}$', pm):
                QMessageBox.warning(
                    self, "تنبيه",
                    "صيغة شهر الصرف غير صحيحة.\nالرجاء الإدخال بالصيغة: YYYY-MM\nمثال: 2026-06"
                )
                return

        self.run_btn.setEnabled(False)
        self.progress.setValue(0)

        self.worker = CompareWorker(
            self.your_drop.file_path,
            self.their_drop.file_path,
            sec,
            self.out_path.text(),
            payment_month=pm
        )
        self.worker.progress.connect(self._on_progress)
        self.worker.finished.connect(self._on_done)
        self.worker.error.connect(self._on_error)
        self.worker.start()

    def _on_progress(self, v, msg):
        self.progress.setValue(v)
        self.status_label.setText(msg)

    def _on_done(self, path):
        self.run_btn.setEnabled(True)
        self.status_label.setText(f"✅  تم الحفظ: {os.path.basename(path)}")

        msg      = QMessageBox(self)
        msg.setWindowTitle("تم بنجاح ✅")
        msg.setText(f"تمت المقارنة بنجاح!\n\nتم حفظ الملف في:\n{path}")
        msg.setIcon(QMessageBox.Icon.Information)

        open_file_btn   = msg.addButton("📄  فتح الملف",   QMessageBox.ButtonRole.AcceptRole)
        open_folder_btn = msg.addButton("📁  فتح الفولدر", QMessageBox.ButtonRole.ActionRole)
        msg.addButton("إغلاق", QMessageBox.ButtonRole.RejectRole)
        msg.exec()

        clicked = msg.clickedButton()
        folder  = os.path.dirname(path)

        def open_path(p):
            if sys.platform == "win32":
                os.startfile(p)
            else:
                os.system(f'xdg-open "{p}"')

        if clicked == open_file_btn:
            open_path(path)
        elif clicked == open_folder_btn:
            open_path(folder)

    def _on_error(self, err):
        self.run_btn.setEnabled(True)
        self.status_label.setText("❌  حدث خطأ")
        QMessageBox.critical(self, "خطأ في المعالجة", f"حدث خطأ:\n\n{err}")


# ─── Entry ────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
    app.setFont(QFont("Arial", 10))
    pal = QPalette()
    pal.setColor(QPalette.ColorRole.Window,     QColor("#F5F7FA"))
    pal.setColor(QPalette.ColorRole.WindowText, QColor("#1B2A4A"))
    app.setPalette(pal)
    win = MainWindow()
    win.showMaximized()
    sys.exit(app.exec())